"""
This script demonstrates how to work with different data file formats in Python.

Supported Formats:
------------------
1. CSV (Comma-Separated Values)
2. PDF (Portable Document Format)
3. Excel (.xlsx files)

Useful Libraries:
-----------------
- csv: for reading and writing CSV files
- PyPDF2: for reading PDF files
- openpyxl: for working with Excel spreadsheets

Each section below includes a working example that shows how to:
- Read data from files
- Write data to files
- Extract useful information
"""

# ----------- 1. CSV Files -----------
import csv

# Reading from a CSV file
with open("sample_data.csv", encoding='utf-8') as my_file:
    csv_data = csv.reader(my_file)
    lines = list(csv_data)

    print("\n📂 CSV File Content:")
    for line in lines:
        print(line)

# Counting a specific column value (e.g., gender in column index 2)
gender_count = {}
for row in lines[1:]:  # Skip header
    gender = row[2]
    gender_count[gender] = gender_count.get(gender, 0) + 1

print("\n📊 Gender Count:", gender_count)

# Writing to a new CSV file
with open("mycsv.csv", mode='w', newline='') as outfile:
    csv_writer = csv.writer(outfile, delimiter='|')
    csv_writer.writerow([1, 2, 3])
    csv_writer.writerow(["Alex", "Monas", "Luca"])

print("\n✅ CSV written to 'mycsv.csv'")

# ----------- 2. PDF Files -----------
import PyPDF2

# Reading from a PDF file
pdf = PyPDF2.PdfReader("sample.pdf")  # Replace with a real file path
num_pages = len(pdf.pages)

print(f"\n📄 PDF has {num_pages} page(s)")

first_page = pdf.pages[0]
text = first_page.extract_text()

print("\n📝 First Page Text:\n", text.strip())

# ----------- 3. Excel Files (.xlsx) -----------
import openpyxl

# Reading from an Excel file
wb = openpyxl.load_workbook("data.xlsx")  # Replace with a real file path
ws = wb.active

# Calculate average from cells B2 to B5
total = 0
counter = 0
for i in range(2, 6):  # B2 to B5
    value = ws[f"B{i}"].value
    if isinstance(value, (int, float)):
        total += value
        counter += 1

if counter > 0:
    avg = total / counter
    print(f"\n📊 Average Age: {avg}")
else:
    print("\n⚠️ No valid data found for average.")

# Writing to a new Excel file
new_wb = openpyxl.Workbook()
ws_new = new_wb.active

# Writing values to specific cells
ws_new["A1"] = "H"
ws_new["B2"] = "E"
ws_new["C3"] = "L"
ws_new["D4"] = "L"
ws_new["E5"] = "O"

# Appending a row
ws_new.append([1, 2, 3, 4])

# Writing the current timestamp
import datetime
ws_new["A10"] = datetime.datetime.now()

# Saving the new Excel file
new_wb.save("funny.xlsx")
print("\n✅ Excel written to 'funny.xlsx'")
